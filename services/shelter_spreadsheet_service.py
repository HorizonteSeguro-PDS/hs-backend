from dataclasses import dataclass
from io import BytesIO
from uuid import UUID
from zipfile import BadZipFile

from fastapi import HTTPException, status
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import select
from sqlalchemy.orm import Session

from domain.errors.http import ResourceNotFoundError
from domain.inventory.schemas import (
    InventoryMovementCreateRequest,
    ShelterSpreadsheetImportResponse,
)
from domain.models.beneficiary import Beneficiary
from domain.models.resource_category import ResourceCategory
from domain.models.shelter import Shelter
from domain.models.shelter_stay import ShelterStay
from domain.schemas.enums import MovementDirection, MovementReason
from repositories.inventory_item import InventoryItemRepository
from repositories.resource_category import ResourceCategoryRepository
from services.inventory_service import InventoryService


RESOURCE_SHEET = "Recursos"
PEOPLE_SHEET = "Pessoas"
RESOURCE_HEADERS = [
    "categoria",
    "unidade",
    "quantidade",
    "observacoes",
]
PEOPLE_HEADERS = [
    "cpf",
    "nome",
    "idade",
    "vulnerabilidade",
    "observacoes",
    "check_in",
    "check_out",
]
SHEET_HEADER_COLORS = {
    RESOURCE_SHEET: "D9EAD3",
    PEOPLE_SHEET: "D9EAF7",
}
COLUMN_WIDTHS = {
    RESOURCE_SHEET: {
        "A": 28,
        "B": 14,
        "C": 14,
        "D": 40,
    },
    PEOPLE_SHEET: {
        "A": 18,
        "B": 30,
        "C": 10,
        "D": 18,
        "E": 48,
        "F": 24,
        "G": 24,
    },
}
MAX_INVENTORY_QUANTITY = 2_147_483_647


class SpreadsheetValidationError(HTTPException):
    def __init__(self, errors: list[str]) -> None:
        super().__init__(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail={"error": "invalid_spreadsheet", "errors": errors},
        )


@dataclass(frozen=True)
class _ResourceRow:
    row_number: int
    category: ResourceCategory
    target_quantity: int
    notes: str | None


@dataclass(frozen=True)
class _ResourceAdjustment:
    category_id: UUID
    direction: MovementDirection
    quantity: int
    notes: str | None


class ShelterSpreadsheetService:
    def __init__(self, session: Session) -> None:
        self.session = session
        self.items = InventoryItemRepository(session)
        self.categories = ResourceCategoryRepository(session)
        self.inventory = InventoryService(session)

    def build_template(self, *, shelter_id: UUID) -> bytes:
        self._ensure_shelter_exists(shelter_id)
        workbook = self._create_workbook()
        return self._to_bytes(workbook)

    def export_spreadsheet(self, *, shelter_id: UUID) -> bytes:
        self._ensure_shelter_exists(shelter_id)
        workbook = self._create_workbook()
        sheet = workbook[RESOURCE_SHEET]

        for item in self.items.list_for_shelter(shelter_id=shelter_id):
            category = self.categories.get(item.category_id)
            sheet.append(
                [
                    category.name if category is not None else "",
                    category.unit.value if category is not None else "",
                    item.quantity_current,
                    "",
                ]
            )
        self._write_people_rows(workbook[PEOPLE_SHEET], shelter_id=shelter_id)

        return self._to_bytes(workbook)

    def import_spreadsheet(
        self,
        *,
        shelter_id: UUID,
        actor_id: UUID,
        content: bytes,
    ) -> ShelterSpreadsheetImportResponse:
        self._ensure_shelter_exists(shelter_id)
        workbook = self._load_workbook(content)
        resource_rows = self._validate_resource_sheet(workbook)
        adjustments = self._plan_adjustments(
            shelter_id=shelter_id,
            resource_rows=resource_rows,
        )

        for adjustment in adjustments:
            self.inventory.record_movement(
                shelter_id=shelter_id,
                actor_id=actor_id,
                payload=InventoryMovementCreateRequest(
                    category_id=adjustment.category_id,
                    direction=adjustment.direction,
                    quantity=adjustment.quantity,
                    reason=MovementReason.ADJUSTMENT,
                    source="spreadsheet_import",
                    notes=adjustment.notes or "Spreadsheet import adjustment",
                ),
            )

        return ShelterSpreadsheetImportResponse(
            shelter_id=shelter_id,
            resources_imported=len(adjustments),
            people_imported=0,
            people_skipped=True,
            errors=[],
        )

    def _ensure_shelter_exists(self, shelter_id: UUID) -> None:
        if self.session.get(Shelter, shelter_id) is None:
            raise ResourceNotFoundError("shelter not found")

    def _create_workbook(self) -> Workbook:
        workbook = Workbook()
        resource_sheet = workbook.active
        resource_sheet.title = RESOURCE_SHEET
        people_sheet = workbook.create_sheet(PEOPLE_SHEET)
        self._write_headers(resource_sheet, RESOURCE_HEADERS)
        self._write_headers(people_sheet, PEOPLE_HEADERS)
        return workbook

    def _write_headers(self, sheet: Worksheet, headers: list[str]) -> None:
        sheet.append(headers)
        sheet.freeze_panes = "A2"
        fill = PatternFill(
            fill_type="solid",
            fgColor=SHEET_HEADER_COLORS.get(sheet.title, "EAEAEA"),
        )
        sheet.sheet_properties.tabColor = SHEET_HEADER_COLORS.get(sheet.title, "EAEAEA")
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = fill
        for column, width in COLUMN_WIDTHS.get(sheet.title, {}).items():
            sheet.column_dimensions[column].width = width

    def _write_people_rows(self, sheet: Worksheet, *, shelter_id: UUID) -> None:
        stmt = (
            select(Beneficiary, ShelterStay)
            .join(ShelterStay, ShelterStay.beneficiary_id == Beneficiary.id)
            .where(
                ShelterStay.shelter_id == shelter_id,
                ShelterStay.checked_out_at.is_(None),
            )
            .order_by(Beneficiary.name)
        )
        for beneficiary, stay in self.session.execute(stmt):
            sheet.append(
                [
                    beneficiary.cpf,
                    beneficiary.name,
                    beneficiary.age,
                    (
                        beneficiary.vulnerability.value
                        if beneficiary.vulnerability is not None
                        else None
                    ),
                    beneficiary.notes,
                    self._format_datetime(stay.checked_in_at),
                    self._format_datetime(stay.checked_out_at),
                ]
            )

    def _load_workbook(self, content: bytes) -> Workbook:
        try:
            return load_workbook(BytesIO(content), data_only=True)
        except (BadZipFile, OSError, ValueError):
            raise SpreadsheetValidationError(["arquivo .xlsx invalido"])

    def _validate_resource_sheet(self, workbook: Workbook) -> list[_ResourceRow]:
        if RESOURCE_SHEET not in workbook.sheetnames:
            raise SpreadsheetValidationError([f"aba '{RESOURCE_SHEET}' ausente"])

        sheet = workbook[RESOURCE_SHEET]
        errors: list[str] = []
        self._validate_headers(sheet, RESOURCE_HEADERS, errors)
        rows: list[_ResourceRow] = []
        seen_categories: set[UUID] = set()

        for row_number, values in self._iter_data_rows(sheet):
            category = self._resolve_category(row_number, values, errors)
            quantity = self._parse_quantity(row_number, values[2], errors)
            notes = self._normalize_optional_text(values[3])

            if category is None or quantity is None:
                continue
            if category.id in seen_categories:
                errors.append(f"linha {row_number}: categoria duplicada")
                continue

            seen_categories.add(category.id)
            rows.append(
                _ResourceRow(
                    row_number=row_number,
                    category=category,
                    target_quantity=quantity,
                    notes=notes,
                )
            )

        if errors:
            raise SpreadsheetValidationError(errors)
        return rows

    def _validate_headers(
        self, sheet: Worksheet, expected: list[str], errors: list[str]
    ) -> None:
        actual = [
            sheet.cell(row=1, column=i).value for i in range(1, len(expected) + 1)
        ]
        if actual != expected:
            errors.append(
                f"aba '{sheet.title}' deve conter headers: {', '.join(expected)}"
            )

    def _iter_data_rows(self, sheet: Worksheet):
        for row_number in range(2, sheet.max_row + 1):
            values = [
                sheet.cell(row=row_number, column=i).value
                for i in range(1, len(RESOURCE_HEADERS) + 1)
            ]
            if all(self._is_blank(value) for value in values):
                continue
            yield row_number, values

    def _resolve_category(
        self, row_number: int, values: list[object], errors: list[str]
    ) -> ResourceCategory | None:
        raw_name = self._normalize_optional_text(values[0])
        if raw_name is None:
            errors.append(f"linha {row_number}: categoria obrigatoria")
            return None

        category = self.categories.get_by_name(raw_name)
        if category is None:
            errors.append(f"linha {row_number}: categoria '{raw_name}' nao encontrada")
        return category

    def _parse_quantity(
        self, row_number: int, value: object, errors: list[str]
    ) -> int | None:
        if self._is_blank(value) or isinstance(value, bool):
            errors.append(f"linha {row_number}: quantidade obrigatoria")
            return None

        if isinstance(value, int):
            quantity = value
        elif isinstance(value, float) and value.is_integer():
            quantity = int(value)
        elif isinstance(value, str):
            text = value.strip()
            try:
                quantity = int(text)
            except ValueError:
                errors.append(f"linha {row_number}: quantidade deve ser inteiro")
                return None
        else:
            errors.append(f"linha {row_number}: quantidade deve ser inteiro")
            return None

        if quantity < 0:
            errors.append(f"linha {row_number}: quantidade nao pode ser negativa")
            return None
        if quantity > MAX_INVENTORY_QUANTITY:
            errors.append(
                f"linha {row_number}: quantidade deve ser no maximo "
                f"{MAX_INVENTORY_QUANTITY}"
            )
            return None
        return quantity

    def _plan_adjustments(
        self,
        *,
        shelter_id: UUID,
        resource_rows: list[_ResourceRow],
    ) -> list[_ResourceAdjustment]:
        adjustments: list[_ResourceAdjustment] = []

        for row in resource_rows:
            item = self.items.get_for_shelter_category(
                shelter_id=shelter_id,
                category_id=row.category.id,
            )
            current_quantity = item.quantity_current if item is not None else 0
            delta = row.target_quantity - current_quantity
            if delta == 0:
                continue

            adjustments.append(
                _ResourceAdjustment(
                    category_id=row.category.id,
                    direction=(
                        MovementDirection.IN if delta > 0 else MovementDirection.OUT
                    ),
                    quantity=abs(delta),
                    notes=row.notes,
                )
            )

        return adjustments

    def _normalize_optional_text(self, value: object) -> str | None:
        if self._is_blank(value):
            return None
        return str(value).strip()

    def _is_blank(self, value: object) -> bool:
        return value is None or (isinstance(value, str) and value.strip() == "")

    def _format_datetime(self, value: object) -> str | None:
        if value is None:
            return None
        return str(value)

    def _to_bytes(self, workbook: Workbook) -> bytes:
        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()
