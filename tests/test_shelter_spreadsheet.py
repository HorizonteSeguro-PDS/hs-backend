import uuid
from datetime import datetime, timedelta, timezone
from io import BytesIO
from unittest.mock import MagicMock

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session

import controllers.shelter_spreadsheet as spreadsheet_controller
import scripts.seed as seed_script
from dependencies.session import get_session
from domain.models.beneficiary import Beneficiary
from domain.models.inventory_item import InventoryItem
from domain.models.inventory_movement import InventoryMovement
from domain.models.resource_category import ResourceCategory
from domain.models.shelter import Shelter
from domain.models.shelter_stay import ShelterStay
from domain.schemas.enums import (
    BrazilianState,
    LotCategory,
    MovementDirection,
    MovementReason,
    ResourceUnit,
    ShelterStatus,
    ShelterType,
    VulnerabilityType,
)
from main import app
from services.shelter_spreadsheet_service import (
    MAX_INVENTORY_QUANTITY,
    PEOPLE_HEADERS,
    PEOPLE_SHEET,
    RESOURCE_HEADERS,
    RESOURCE_SHEET,
    ShelterSpreadsheetService,
    SpreadsheetValidationError,
)
from tests.conftest import auth_headers


def _setup_session() -> Session:
    engine = create_engine("sqlite:///:memory:")
    for table in (
        Shelter.__table__,
        Beneficiary.__table__,
        ShelterStay.__table__,
        ResourceCategory.__table__,
        InventoryItem.__table__,
        InventoryMovement.__table__,
    ):
        table.create(engine)
    return Session(engine)


def _seed_shelter(
    session: Session,
    shelter_id: uuid.UUID | None = None,
    name: str = "Abrigo Central",
) -> Shelter:
    shelter = Shelter(
        id=shelter_id or uuid.uuid4(),
        responsible_user_id=uuid.uuid4(),
        created_by=uuid.uuid4(),
        name=name,
        address="Rua Principal, 100",
        city="Sao Paulo",
        state=BrazilianState.SP,
        capacity=100,
        occupation=25,
        shelter_type=ShelterType.INSTITUTIONAL,
        status=ShelterStatus.ACTIVE,
        verified=True,
    )
    session.add(shelter)
    session.commit()
    return shelter


def _seed_category(
    session: Session,
    name: str = "cobertor",
    unit: ResourceUnit = ResourceUnit.UNIDADE,
) -> ResourceCategory:
    category = ResourceCategory(
        id=uuid.uuid4(),
        name=name,
        unit=unit,
        lot_category=LotCategory.OPERACAO,
        description=None,
    )
    session.add(category)
    session.commit()
    return category


def _seed_item(
    session: Session,
    *,
    shelter_id: uuid.UUID,
    category_id: uuid.UUID,
    quantity: int,
) -> InventoryItem:
    item = InventoryItem(
        id=uuid.uuid4(),
        shelter_id=shelter_id,
        category_id=category_id,
        quantity_current=quantity,
    )
    session.add(item)
    session.commit()
    return item


def _seed_person(
    session: Session,
    *,
    shelter_id: uuid.UUID,
    name: str,
    cpf: str,
    checked_out_at: datetime | None = None,
) -> Beneficiary:
    beneficiary = Beneficiary(
        id=uuid.uuid4(),
        cpf=cpf,
        name=name,
        age=30,
        vulnerability=VulnerabilityType.NONE,
        notes="observacao",
    )
    session.add(beneficiary)
    session.flush()
    session.add(
        ShelterStay(
            id=uuid.uuid4(),
            beneficiary_id=beneficiary.id,
            shelter_id=shelter_id,
            checked_in_at=datetime(2026, 1, 1, tzinfo=timezone.utc),
            checked_out_at=checked_out_at,
        )
    )
    session.commit()
    return beneficiary


def _load(content: bytes) -> Workbook:
    return load_workbook(BytesIO(content), data_only=True)


def _workbook_bytes(
    rows: list[list[object]], people_rows: list[list[object]] | None = None
) -> bytes:
    workbook = Workbook()
    resources = workbook.active
    resources.title = RESOURCE_SHEET
    resources.append(RESOURCE_HEADERS)
    for row in rows:
        resources.append(row)
    people = workbook.create_sheet(PEOPLE_SHEET)
    people.append(PEOPLE_HEADERS)
    for row in people_rows or []:
        people.append(row)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _workbook_without_resource_sheet() -> bytes:
    workbook = Workbook()
    workbook.active.title = PEOPLE_SHEET
    workbook[PEOPLE_SHEET].append(PEOPLE_HEADERS)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _workbook_with_resource_headers(headers: list[str]) -> bytes:
    workbook = Workbook()
    resources = workbook.active
    resources.title = RESOURCE_SHEET
    resources.append(headers)
    workbook.create_sheet(PEOPLE_SHEET).append(PEOPLE_HEADERS)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _session_override(mock: MagicMock | None = None):
    def override():
        yield mock or MagicMock()

    return override


class TestSeedSpreadsheetData:
    def test_demo_shelter_id_belongs_to_demo_shelter_spec(self):
        demo_specs = [
            spec
            for spec in seed_script.SHELTERS_SPEC
            if spec["name"] == seed_script.DEMO_FULL_SHELTER_NAME
        ]
        misplaced_specs = [
            spec
            for spec in seed_script.SHELTERS_SPEC
            if spec["name"] != seed_script.DEMO_FULL_SHELTER_NAME
            and spec.get("id") == seed_script.DEMO_FULL_SHELTER_ID
        ]

        assert len(demo_specs) == 1
        assert demo_specs[0]["id"] == seed_script.DEMO_FULL_SHELTER_ID
        assert misplaced_specs == []

    def test_seed_demo_full_shelter_open_stay_is_scoped_to_shelter(self):
        with _setup_session() as session:
            demo_shelter = _seed_shelter(
                session,
                shelter_id=seed_script.DEMO_FULL_SHELTER_ID,
                name=seed_script.DEMO_FULL_SHELTER_NAME,
            )
            other_shelter = _seed_shelter(session, name="Outro Abrigo")
            seeded_person = seed_script.BENEFICIARIES_SPEC[0]
            beneficiary = _seed_person(
                session,
                shelter_id=other_shelter.id,
                name=seeded_person["name"],
                cpf=seeded_person["cpf"],
            )
            categories = [
                _seed_category(session, name, unit)
                for name, unit, _lot, _description in seed_script.CATEGORIES_SPEC
            ]

            seed_script.seed_demo_full_shelter(
                session,
                shelters=[demo_shelter, other_shelter],
                categories=categories,
                actor_id=uuid.uuid4(),
            )
            session.commit()

            demo_open_stays = session.query(ShelterStay).filter_by(
                beneficiary_id=beneficiary.id,
                shelter_id=demo_shelter.id,
                checked_out_at=None,
            )
            other_open_stays = session.query(ShelterStay).filter_by(
                beneficiary_id=beneficiary.id,
                shelter_id=other_shelter.id,
                checked_out_at=None,
            )
            assert demo_open_stays.count() == 1
            assert other_open_stays.count() == 1


class TestShelterSpreadsheetService:
    def test_format_datetime_serializes_naive_and_timezone_aware_values(self):
        service = ShelterSpreadsheetService(MagicMock())

        assert (
            service._format_datetime(datetime(2026, 1, 2, 3, 4, 5))
            == "2026-01-02 03:04:05"
        )
        assert (
            service._format_datetime(
                datetime(
                    2026,
                    1,
                    2,
                    0,
                    4,
                    5,
                    tzinfo=timezone(timedelta(hours=-3)),
                )
            )
            == "2026-01-02 03:04:05"
        )
        assert service._format_datetime("manual") == "manual"

    def test_template_has_expected_sheets_and_headers(self):
        with _setup_session() as session:
            _seed_shelter(session)

            workbook = _load(ShelterSpreadsheetService(session).build_template())

            resource_headers = [cell.value for cell in workbook[RESOURCE_SHEET][1]]
            assert workbook.sheetnames == [RESOURCE_SHEET, PEOPLE_SHEET]
            assert resource_headers == RESOURCE_HEADERS
            assert [cell.value for cell in workbook[PEOPLE_SHEET][1]] == PEOPLE_HEADERS
            assert "category_id" not in resource_headers
            assert workbook[RESOURCE_SHEET].max_row == 1
            assert workbook[PEOPLE_SHEET].max_row == 1

    def test_template_applies_basic_readability_styles(self):
        with _setup_session() as session:
            _seed_shelter(session)

            workbook = _load(ShelterSpreadsheetService(session).build_template())
            resources = workbook[RESOURCE_SHEET]
            people = workbook[PEOPLE_SHEET]

            assert resources.freeze_panes == "A2"
            assert people.freeze_panes == "A2"
            assert resources["A1"].font.bold is True
            assert people["A1"].font.bold is True
            assert resources["A1"].fill.fill_type == "solid"
            assert people["A1"].fill.fill_type == "solid"
            assert resources.column_dimensions["A"].width == 28
            assert resources.column_dimensions["D"].width == 40
            assert people.column_dimensions["B"].width == 30
            assert people.column_dimensions["E"].width == 48

    def test_export_includes_current_resources(self):
        with _setup_session() as session:
            shelter = _seed_shelter(session)
            category = _seed_category(session, "agua", ResourceUnit.L)
            _seed_item(
                session,
                shelter_id=shelter.id,
                category_id=category.id,
                quantity=12,
            )

            workbook = _load(
                ShelterSpreadsheetService(session).export_spreadsheet(
                    shelter_id=shelter.id
                )
            )
            row = [cell.value for cell in workbook[RESOURCE_SHEET][2]]

            assert row == ["agua", "L", 12, None]
            assert PEOPLE_SHEET in workbook.sheetnames

    def test_export_includes_active_people_only(self):
        with _setup_session() as session:
            shelter = _seed_shelter(session)
            other_shelter = _seed_shelter(session)
            active = _seed_person(
                session,
                shelter_id=shelter.id,
                name="Maria",
                cpf="000.000.000-01",
            )
            _seed_person(
                session,
                shelter_id=shelter.id,
                name="Joao",
                cpf="000.000.000-02",
                checked_out_at=datetime(2026, 1, 2, tzinfo=timezone.utc),
            )
            _seed_person(
                session,
                shelter_id=other_shelter.id,
                name="Ze Outro",
                cpf="000.000.000-03",
            )

            workbook = _load(
                ShelterSpreadsheetService(session).export_spreadsheet(
                    shelter_id=shelter.id
                )
            )
            rows = list(workbook[PEOPLE_SHEET].iter_rows(min_row=2, values_only=True))

            assert rows == [
                (
                    active.cpf,
                    active.name,
                    active.age,
                    active.vulnerability.value,
                    active.notes,
                    "2026-01-01 00:00:00",
                    None,
                )
            ]

    def test_import_creates_adjustment_movements(self):
        with _setup_session() as session:
            shelter = _seed_shelter(session)
            category = _seed_category(session)
            _seed_item(
                session,
                shelter_id=shelter.id,
                category_id=category.id,
                quantity=10,
            )
            actor_id = uuid.uuid4()
            content = _workbook_bytes(
                [[category.name, category.unit.value, 15, "contagem"]]
            )

            result = ShelterSpreadsheetService(session).import_spreadsheet(
                shelter_id=shelter.id,
                actor_id=actor_id,
                content=content,
            )
            session.commit()

            movement = session.query(InventoryMovement).one()
            item = session.query(InventoryItem).one()
            assert result.resources_imported == 1
            assert result.people_imported == 0
            assert result.people_skipped is True
            assert movement.direction == MovementDirection.IN
            assert movement.reason == MovementReason.ADJUSTMENT
            assert movement.quantity == 5
            assert movement.created_by == actor_id
            assert item.quantity_current == 15

    def test_import_creates_out_adjustment_when_quantity_decreases(self):
        with _setup_session() as session:
            shelter = _seed_shelter(session)
            category = _seed_category(session)
            _seed_item(
                session,
                shelter_id=shelter.id,
                category_id=category.id,
                quantity=10,
            )
            content = _workbook_bytes([[category.name, category.unit.value, 4, None]])

            result = ShelterSpreadsheetService(session).import_spreadsheet(
                shelter_id=shelter.id,
                actor_id=uuid.uuid4(),
                content=content,
            )
            session.commit()

            movement = session.query(InventoryMovement).one()
            item = session.query(InventoryItem).one()
            assert result.resources_imported == 1
            assert movement.direction == MovementDirection.OUT
            assert movement.reason == MovementReason.ADJUSTMENT
            assert movement.quantity == 6
            assert item.quantity_current == 4

    def test_import_skips_unchanged_quantities(self):
        with _setup_session() as session:
            shelter = _seed_shelter(session)
            category = _seed_category(session)
            _seed_item(
                session,
                shelter_id=shelter.id,
                category_id=category.id,
                quantity=10,
            )
            content = _workbook_bytes([[category.name, category.unit.value, 10, None]])

            result = ShelterSpreadsheetService(session).import_spreadsheet(
                shelter_id=shelter.id,
                actor_id=uuid.uuid4(),
                content=content,
            )

            assert result.resources_imported == 0
            assert session.query(InventoryMovement).count() == 0
            assert session.query(InventoryItem).one().quantity_current == 10

    def test_import_creates_inventory_item_for_existing_category_without_stock(self):
        with _setup_session() as session:
            shelter = _seed_shelter(session)
            category = _seed_category(session, "colchao")
            content = _workbook_bytes([[category.name, category.unit.value, 7, None]])

            result = ShelterSpreadsheetService(session).import_spreadsheet(
                shelter_id=shelter.id,
                actor_id=uuid.uuid4(),
                content=content,
            )
            session.commit()

            item = session.query(InventoryItem).one()
            movement = session.query(InventoryMovement).one()
            assert result.resources_imported == 1
            assert item.category_id == category.id
            assert item.quantity_current == 7
            assert movement.direction == MovementDirection.IN
            assert movement.quantity == 7

    def test_import_validates_all_rows_before_applying(self):
        with _setup_session() as session:
            shelter = _seed_shelter(session)
            category = _seed_category(session)
            _seed_item(
                session,
                shelter_id=shelter.id,
                category_id=category.id,
                quantity=10,
            )
            content = _workbook_bytes(
                [
                    [category.name, category.unit.value, 20, None],
                    ["fantasma", "unidade", 5, None],
                ]
            )

            with pytest.raises(SpreadsheetValidationError) as exc:
                ShelterSpreadsheetService(session).import_spreadsheet(
                    shelter_id=shelter.id,
                    actor_id=uuid.uuid4(),
                    content=content,
                )

            assert (
                "linha 3: categoria 'fantasma' nao encontrada"
                in exc.value.detail["errors"]
            )
            item = session.query(InventoryItem).one()
            assert session.query(InventoryMovement).count() == 0
            assert item.quantity_current == 10

    @pytest.mark.parametrize(
        ("quantity", "expected_error"),
        [
            (-1, "linha 2: quantidade nao pode ser negativa"),
            (1.5, "linha 2: quantidade deve ser inteiro"),
            ("abc", "linha 2: quantidade deve ser inteiro"),
            ("", "linha 2: quantidade obrigatoria"),
            (True, "linha 2: quantidade obrigatoria"),
        ],
    )
    def test_import_rejects_invalid_quantities(self, quantity, expected_error):
        with _setup_session() as session:
            shelter = _seed_shelter(session)
            category = _seed_category(session)
            content = _workbook_bytes(
                [[category.name, category.unit.value, quantity, None]]
            )

            with pytest.raises(SpreadsheetValidationError) as exc:
                ShelterSpreadsheetService(session).import_spreadsheet(
                    shelter_id=shelter.id,
                    actor_id=uuid.uuid4(),
                    content=content,
                )

            assert expected_error in exc.value.detail["errors"]
            assert session.query(InventoryMovement).count() == 0

    def test_import_rejects_duplicate_category_rows(self):
        with _setup_session() as session:
            shelter = _seed_shelter(session)
            category = _seed_category(session)
            content = _workbook_bytes(
                [
                    [category.name, category.unit.value, 10, None],
                    [category.name, category.unit.value, 20, None],
                ]
            )

            with pytest.raises(SpreadsheetValidationError) as exc:
                ShelterSpreadsheetService(session).import_spreadsheet(
                    shelter_id=shelter.id,
                    actor_id=uuid.uuid4(),
                    content=content,
                )

            assert "linha 3: categoria duplicada" in exc.value.detail["errors"]
            assert session.query(InventoryMovement).count() == 0

    def test_import_rejects_missing_resource_sheet(self):
        with _setup_session() as session:
            shelter = _seed_shelter(session)

            with pytest.raises(SpreadsheetValidationError) as exc:
                ShelterSpreadsheetService(session).import_spreadsheet(
                    shelter_id=shelter.id,
                    actor_id=uuid.uuid4(),
                    content=_workbook_without_resource_sheet(),
                )

            assert "aba 'Recursos' ausente" in exc.value.detail["errors"]

    def test_import_rejects_unexpected_resource_headers(self):
        with _setup_session() as session:
            shelter = _seed_shelter(session)

            with pytest.raises(SpreadsheetValidationError) as exc:
                ShelterSpreadsheetService(session).import_spreadsheet(
                    shelter_id=shelter.id,
                    actor_id=uuid.uuid4(),
                    content=_workbook_with_resource_headers(["category_id"]),
                )

            assert (
                "aba 'Recursos' deve conter headers: categoria, unidade, "
                "quantidade, observacoes"
            ) in exc.value.detail["errors"]

    def test_import_rejects_quantity_above_database_integer_limit(self):
        with _setup_session() as session:
            shelter = _seed_shelter(session)
            category = _seed_category(session, "kit_medico_basico")
            _seed_item(
                session,
                shelter_id=shelter.id,
                category_id=category.id,
                quantity=45,
            )
            content = _workbook_bytes(
                [[category.name, category.unit.value, 55_555_555_555, None]]
            )

            with pytest.raises(SpreadsheetValidationError) as exc:
                ShelterSpreadsheetService(session).import_spreadsheet(
                    shelter_id=shelter.id,
                    actor_id=uuid.uuid4(),
                    content=content,
                )

            assert (
                "linha 2: quantidade deve ser no maximo 2147483647"
                in (exc.value.detail["errors"])
            )
            assert MAX_INVENTORY_QUANTITY == 2_147_483_647
            item = session.query(InventoryItem).one()
            assert session.query(InventoryMovement).count() == 0
            assert item.quantity_current == 45

    def test_import_rejects_invalid_file(self):
        with _setup_session() as session:
            shelter = _seed_shelter(session)

            with pytest.raises(SpreadsheetValidationError) as exc:
                ShelterSpreadsheetService(session).import_spreadsheet(
                    shelter_id=shelter.id,
                    actor_id=uuid.uuid4(),
                    content=b"not an xlsx",
                )

            assert exc.value.status_code == 400

    def test_people_sheet_is_ignored_on_import(self):
        with _setup_session() as session:
            shelter = _seed_shelter(session)
            content = _workbook_bytes(
                [],
                people_rows=[
                    [
                        "000.000.000-00",
                        "Maria",
                        30,
                        "none",
                        "sem persistencia",
                        None,
                        None,
                    ]
                ],
            )

            result = ShelterSpreadsheetService(session).import_spreadsheet(
                shelter_id=shelter.id,
                actor_id=uuid.uuid4(),
                content=content,
            )

            assert result.resources_imported == 0
            assert result.people_imported == 0
            assert result.people_skipped is True
            assert session.query(InventoryMovement).count() == 0


class TestShelterSpreadsheetController:
    def setup_method(self):
        app.dependency_overrides = {}

    def teardown_method(self):
        app.dependency_overrides = {}

    def test_template_download_returns_xlsx(self, monkeypatch):
        content = _workbook_bytes([])

        class FakeShelterSpreadsheetService:
            def __init__(self, session):
                self.session = session

            def build_template(self):
                return content

        monkeypatch.setattr(
            spreadsheet_controller,
            "ShelterSpreadsheetService",
            FakeShelterSpreadsheetService,
        )
        app.dependency_overrides[get_session] = _session_override()

        response = TestClient(app).get(
            "/shelters/spreadsheet/template",
            headers=auth_headers("shelter_manager"),
        )

        workbook = _load(response.content)
        assert response.status_code == 200
        assert (
            response.headers["content-type"] == spreadsheet_controller.XLSX_MEDIA_TYPE
        )
        assert "attachment;" in response.headers["content-disposition"]
        assert (
            "shelter-spreadsheet-template.xlsx"
            in response.headers["content-disposition"]
        )
        assert [cell.value for cell in workbook[RESOURCE_SHEET][1]] == RESOURCE_HEADERS
        assert [cell.value for cell in workbook[PEOPLE_SHEET][1]] == PEOPLE_HEADERS

    def test_export_download_returns_xlsx(self, monkeypatch):
        shelter_id = uuid.uuid4()
        content = _workbook_bytes([["agua", "L", 12, None]])

        class FakeShelterSpreadsheetService:
            def __init__(self, session):
                self.session = session

            def export_spreadsheet(self, *, shelter_id):
                return content

        monkeypatch.setattr(
            spreadsheet_controller,
            "ShelterSpreadsheetService",
            FakeShelterSpreadsheetService,
        )
        app.dependency_overrides[get_session] = _session_override()

        response = TestClient(app).get(
            f"/shelters/{shelter_id}/spreadsheet/export",
            headers=auth_headers("dev"),
        )

        workbook = _load(response.content)
        assert response.status_code == 200
        assert (
            response.headers["content-type"] == spreadsheet_controller.XLSX_MEDIA_TYPE
        )
        assert (
            f"shelter-{shelter_id}-export.xlsx"
            in response.headers["content-disposition"]
        )
        assert [cell.value for cell in workbook[RESOURCE_SHEET][2]] == [
            "agua",
            "L",
            12,
            None,
        ]

    def test_import_accepts_xlsx_upload(self, monkeypatch):
        session = MagicMock()
        shelter_id = uuid.uuid4()
        user_id = uuid.uuid4()
        content = _workbook_bytes([])
        captured = {}

        class FakeShelterSpreadsheetService:
            def __init__(self, session):
                self.session = session

            def import_spreadsheet(self, *, shelter_id, actor_id, content):
                captured["shelter_id"] = shelter_id
                captured["actor_id"] = actor_id
                captured["content"] = content
                return {
                    "shelter_id": shelter_id,
                    "resources_imported": 0,
                    "people_imported": 0,
                    "people_skipped": True,
                    "errors": [],
                }

        monkeypatch.setattr(
            spreadsheet_controller,
            "ShelterSpreadsheetService",
            FakeShelterSpreadsheetService,
        )
        app.dependency_overrides[get_session] = _session_override(session)

        response = TestClient(app).post(
            f"/shelters/{shelter_id}/spreadsheet/import",
            files={
                "file": (
                    "inventory.xlsx",
                    content,
                    spreadsheet_controller.XLSX_MEDIA_TYPE,
                )
            },
            headers=auth_headers("crisis_manager", user_id=str(user_id)),
        )

        assert response.status_code == 200
        assert response.json()["people_skipped"] is True
        assert captured["shelter_id"] == shelter_id
        assert captured["actor_id"] == user_id
        assert captured["content"] == content
        session.commit.assert_called_once()

    def test_import_validation_error_does_not_commit(self, monkeypatch):
        session = MagicMock()
        shelter_id = uuid.uuid4()

        class FakeShelterSpreadsheetService:
            def __init__(self, session):
                self.session = session

            def import_spreadsheet(self, *, shelter_id, actor_id, content):
                raise SpreadsheetValidationError(["linha 2: quantidade obrigatoria"])

        monkeypatch.setattr(
            spreadsheet_controller,
            "ShelterSpreadsheetService",
            FakeShelterSpreadsheetService,
        )
        app.dependency_overrides[get_session] = _session_override(session)

        response = TestClient(app).post(
            f"/shelters/{shelter_id}/spreadsheet/import",
            files={
                "file": (
                    "inventory.xlsx",
                    _workbook_bytes([]),
                    spreadsheet_controller.XLSX_MEDIA_TYPE,
                )
            },
            headers=auth_headers("shelter_manager"),
        )

        assert response.status_code == 400
        assert response.json()["detail"]["error"] == "invalid_spreadsheet"
        session.commit.assert_not_called()
